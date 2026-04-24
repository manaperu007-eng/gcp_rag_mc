##############################################################################
# backend/ingestion/processors/excel_processor.py
# Extract text from .xlsx / .xls files using openpyxl
##############################################################################
from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """
    Converts Excel workbooks into a list of page-like dicts, one per sheet.

    Output page dict:
      - ``text``         : tab-separated cell values, rows separated by newline
      - ``page_number``  : 1-based sheet index
      - ``section_title``: sheet name
    """

    def extract(self, file_path: str) -> List[Dict[str, Any]]:
        """Read an Excel file and return one "page" per worksheet."""
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel processing. "
                "Install it with: pip install openpyxl"
            ) from exc

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            # Fall back to xlrd for legacy .xls files
            logger.warning("openpyxl failed (%s), trying xlrd for .xls", exc)
            return self._extract_xls(file_path)

        pages: List[Dict[str, Any]] = []
        for idx, sheet_name in enumerate(workbook.sheetnames, start=1):
            ws = workbook[sheet_name]
            rows_text = self._sheet_to_text(ws)
            if rows_text.strip():
                pages.append({
                    "text": rows_text,
                    "page_number": idx,
                    "section_title": sheet_name,
                })

        workbook.close()

        if not pages:
            logger.warning("Excel file %s produced no text", file_path)

        return pages

    # ── Internal ──────────────────────────────────────────────────────────

    @staticmethod
    def _sheet_to_text(ws) -> str:
        """Convert a worksheet to a flat text string."""
        lines: List[str] = []
        for row in ws.iter_rows(values_only=True):
            cell_values = [
                str(cell).strip() if cell is not None else ""
                for cell in row
            ]
            # Skip fully empty rows
            if any(v for v in cell_values):
                lines.append("\t".join(cell_values))
        return "\n".join(lines)

    def _extract_xls(self, file_path: str) -> List[Dict[str, Any]]:
        """Legacy .xls fallback using xlrd."""
        try:
            import xlrd  # noqa: PLC0415
        except ImportError as exc:
            raise ImportError(
                "xlrd is required for legacy .xls files. "
                "Install it with: pip install xlrd"
            ) from exc

        workbook = xlrd.open_workbook(file_path)
        pages: List[Dict[str, Any]] = []
        for idx, sheet in enumerate(workbook.sheets(), start=1):
            rows_text = self._xls_sheet_to_text(sheet)
            if rows_text.strip():
                pages.append({
                    "text": rows_text,
                    "page_number": idx,
                    "section_title": sheet.name,
                })
        return pages

    @staticmethod
    def _xls_sheet_to_text(sheet) -> str:
        lines: List[str] = []
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
            if any(c for c in cells):
                lines.append("\t".join(cells))
        return "\n".join(lines)
